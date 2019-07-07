from django.shortcuts import render, redirect
from .forms import UplodForm
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from .models import IPC_Codes, Inventions, Excel_Files
import xlwt
import os



def excel_upload(request):
    if request.method == 'POST' and request.FILES['file']:
        myfile = request.FILES['file']
        fs = FileSystemStorage()
        temp=myfile.name.split('.')
        fileexactname=temp[0]
        wb = load_workbook(myfile)
        sheet = wb.active
        if sheet.max_column==2:
            # IPC Classification Start
            if not os.path.isfile('media/'+myfile.name):
                for i in range(1, sheet.max_row):
                    newrecord=IPC_Codes(code= sheet.cell(row=i, column=1).value, title=sheet.cell(row=i, column=2).value)
                    existrecord=IPC_Codes.objects.filter(code__exact=sheet.cell(row=i, column=1).value)
                    if existrecord:
                        if existrecord[0].code!=sheet.cell(row=i, column=1).value:
                            newrecord.save()
                    else:
                        newrecord.save()
                filename = fs.save(myfile.name, myfile)
                uploaded_file_url = fs.url(filename)
                        #print(i, sheet.cell(row=i, column=2).value)
                return redirect('')
            else:
                return render(request, 'p_excel/errorpage.html', {
                    'contect': "This file Uploaded befor!"
                })

            # IPC Classification End
        elif sheet.max_column == 6:
            # Inventions Start
            if not os.path.isfile('media/' + myfile.name):
                existrecord = Inventions.objects.filter(year_week__exact=myfile.name)
                if not existrecord:
                    for i in range(1, sheet.max_row):
                        newrecord=Inventions(name= sheet.cell(row=i, column=1).value, link= sheet.cell(row=i, column=1).hyperlink.target, title=sheet.cell(row=i, column=2).value, untype=sheet.cell(row=i, column=3).value, patent=sheet.cell(row=i, column=4).value, classification=sheet.cell(row=i, column=5).value, inventor=sheet.cell(row=i, column=6).value, year_week=fileexactname)
                        print(newrecord);
                        newrecord.save()
                    filename = fs.save(myfile.name, myfile)
                IPC=IPC_Codes.objects.all()
                for row in IPC:
                    export_excel_file(row.code, row.title, fileexactname)

                files = Excel_Files.objects.filter(year_week__exact=fileexactname)
                return render(request, 'p_excel/download.html', {
                    'files': files
                })
            else:
                # Read list of files
                files=Excel_Files.objects.filter(year_week__exact=fileexactname)

                #DOWNLOADEDFILES={}
                #FOR F IN FILES:
                    #DF=F.YEAR_WEEK+'_'+F.CLASSIFICATION[:3]+'.XLS'
                    #DOWNLOADEDFILES[DF]=F.TITLE
                return render(request, 'p_excel/download.html', {
                    'files': files
                })
            # Inventions End
        else:
            return render(request, 'p_excel/errorpage.html', {
                'contect': "Uploaded file is not correct!"
            })
    else:
        form=UplodForm()
    context= {'form': form}
    return render(request,'p_excel/upload.html', context)

def download(request):
    return render(request,'p_excel/download.html')

def export_excel_file(code1, title1, o_filename):
    Inv = Inventions.objects.filter(classification__startswith=code1)
    if Inv:
        row_num = 1
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('ResultSet')
        filename1 = o_filename + '_' + code1 + '.xls'
        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        ws.col(0).width = 16 * 256
        ws.col(1).width = 255 * 256
        ws.col(2).width = 6 * 256
        ws.col(3).width = 15 * 256
        ws.col(4).width = 13 * 256
        ws.col(5).width = 102 * 256
        ws.write(0, 0, 'ID', font_style)
        ws.write(0, 1, 'Title', font_style)
        ws.write(0, 2, 'Kind', font_style)
        ws.write(0, 3, 'AppI.No', font_style)
        ws.write(0, 4, 'IPC', font_style)
        ws.write(0, 5, 'Applicant', font_style)
        font_style.font.bold = False
        for row in Inv:
            formula = 'HYPERLINK("{}", "{}")'.format(row.link, row.name)
            ws.write(row_num, 0, xlwt.Formula(formula), font_style)
            ws.write(row_num, 1, row.title, font_style)
            ws.write(row_num, 2, row.untype, font_style)
            ws.write(row_num, 3, row.patent, font_style)
            ws.write(row_num, 4, row.classification, font_style)
            ws.write(row_num, 5, row.inventor, font_style)
            row_num += 1

        wb.save('media/'+filename1)
        recex=Excel_Files.objects.filter(filename__exact=filename1)
        if not recex:
            rec=Excel_Files(code=code1, title= title1, filename=filename1, year_week=o_filename)
            rec.save()
